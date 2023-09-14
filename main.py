# create by TurtleOfOceans 06/09/2023
import os
import openpyxl
import pandas as pd
from tqdm import tqdm


def readExcelWithPandas(input_dir, sheet=0):
    """
    read excel file with pandas lib
    Args:
        input_dir: excel director
        sheet: specify sheet to read data
    return:
        excel_data: excel data
    """
    HEADER_LINE = 3
    excel_data = pd.read_excel(input_dir, sheet, header=HEADER_LINE)
    return excel_data


def getDetailSheet(input_dir, sheet=0):
    """
    read excel file with pandas lib
    Args:
        input_dir: excel director
        sheet: specify sheet to read data
    return:
        detail_data: excel data
    """
    excel_data_total = pd.read_excel(input_dir, sheet, header=None)
    detail_data = excel_data_total.iloc[:3]
    return detail_data


def readExcelWithOpenpyxl(input_dir, sheet=0):
    """
    read excel file with openpyxl lib
    Args:
        input_dir: excel director
        sheet: specify sheet to read data
    return:
        work_sheet: sheet data
    """
    work_book = openpyxl.load_workbook(input_dir, data_only=True)
    work_sheet = work_book.worksheets[sheet]
    return work_sheet


def getListHeader(excel_data):
    """
    get list header
    Args:
        excel_data: excel data
    return:
        header_data: header data
    """
    header = list(excel_data.columns.values)
    return header


def getListDataColumn(excel_data, title_name):
    """
    get list data in column
    Args:
        excel_data: excel_data
        title_name: title column
    return:
        filter_data: filter data
    """
    filter_data = []
    column_data = excel_data[title_name]
    for cell in column_data:
        if cell not in filter_data:
            filter_data.append(cell)
    return filter_data


def createMultiExcelFile(output_dir, name, data_frame):
    """
    write data to excel file
    Args:
        output_dir: out put directory
        name: sheet name
        data_frame: data of sheet
    """
    file_name = os.path.join(output_dir, f'{name}.xlsx')
    data_frame.to_excel(file_name, name, index=False)


def splitTableDataToMultiFile(
        excel_data,
        filter_data,
        filter_title,
        output_dir):
    """
    split table data
    Args:
        excel_data: excel_data
        filter_data: filter column
        filter_title: specify column filter
        output_dir: specify directory out put
    """
    rows = len(excel_data.axes[0])
    total = len(filter_data)
    progresBar = tqdm(range(total), desc="Create Fille...")
    file = 0
    for filter in filter_data:
        tmp_table = []
        for i in range(0, rows):
            row_data = excel_data.iloc[i, :]
            if (row_data[filter_title] == filter):
                tmp_table.append(row_data)

        data_frame = pd.DataFrame(tmp_table)
        createMultiExcelFile(output_dir, filter, data_frame)
        file += 1
        progresBar.update()


def createColumnData(total_row, data):
    """
    split table data
    Args:
        total_row: total row
        data: column data
    return:
        data_array: column data
    """
    data_array = []
    total_data = int(len(data))
    for i in range(0, total_row):
        if i == 0:
            data_array.append(total_data)
        else:
            data_array.append(None)
    return data_array


def splitTableDataToMultiSheet(
        excel_data,
        detail_data,
        filter_data,
        filter_title,
        output_dir,
        sheet):
    """
    split table data
    Args:
        excel_data: excel_data
        detail_data: detail_data
        filter_data: filter column
        filter_title: specify column filter
        output_dir: specify directory out put
    """
    MDVCNT = 2
    SUMDVCNT_Hex = "53E1BB91206CC6B0E1BBA36E6720C49056434E54"
    byte_string = bytes.fromhex(SUMDVCNT_Hex)
    SUMDVCNT = byte_string.decode('utf-8')
    START_ROW = 3
    rows = len(excel_data.axes[0])
    total = len(filter_data)
    file_name = os.path.join(output_dir, f'{sheet}.xlsx')
    writer = pd.ExcelWriter(file_name)
    progresBar = tqdm(range(total), desc="Create Fille...")
    file = 0
    detail_data.to_excel(writer, "Total", index=False, header=None)
    excel_data.to_excel(writer, "Total", index=False, startrow=START_ROW)
    all_dvcnt_array = []
    total_dvcnt = 0
    for filter in filter_data:
        tmp_table = []
        dvcnt_array = []
        for i in range(0, rows):
            row_data = excel_data.iloc[i, 1:]
            if (row_data[filter_title] == filter):
                tmp_table.append(row_data)
                if row_data.iloc[MDVCNT] not in all_dvcnt_array:
                    all_dvcnt_array.append(row_data.iloc[MDVCNT])
                if row_data.iloc[MDVCNT] not in dvcnt_array:
                    dvcnt_array.append(row_data.iloc[MDVCNT])

        data_frame = pd.DataFrame(tmp_table)
        data_frame[SUMDVCNT] = createColumnData(len(tmp_table), dvcnt_array)
        index = pd.Index(range(1, len(data_frame)+1))
        total_dvcnt += len(dvcnt_array)
        data_frame.index = index
        detail_data.to_excel(writer, filter, index=False, header=None)
        data_frame.to_excel(
            writer,
            filter,
            index_label="STT",
            startrow=START_ROW)

        file += 1
        progresBar.update()

    writer.close()
    if total_dvcnt == len(all_dvcnt_array):
        print("CHECK TOTAL DVCNT: OK")
    else:
        print("CHECK TOTAL DVCNT: NOT OK")


def inputSplitParam():
    """
    ask the user to enter input
    return:
        param: contain data specify by user
    """
    print("==========================")
    input_dir = input("Input (excel file): ")
    output_dir = input("Output (specify folder): ")
    active_sheet = input("Specify sheet active [0, 1, ...]: ")
    filter = input("Specify filter [0, 1, ...]: ")
    print("==========================")
    param = {
        'inputDir': input_dir,
        'outputDir': output_dir,
        'activeSheet': int(active_sheet),
        'filter': int(filter),
    }
    return param


def inputParamForCreateFilter():
    """
    ask the user to enter input
    return:
        param: contain data specify by user
    """
    print("==========================")
    input_dir = input("Input (excel file): ")
    output_dir = input("Output (specify folder): ")
    active_sheet = input("Specify sheet active [0, 1, ...]: ")
    print("==========================")
    param = {
        'inputDir': input_dir,
        'outputDir': output_dir,
        'activeSheet': int(active_sheet),
    }
    return param


def splitDataWithMode(input_dir, output_dir, filter, sheet, mode):
    """
    read file excel and split
    Args:
        input_dir: excel director
        output_dir: specify out put data
        filter: specify filter
        sheet: specify sheet by user
        mode: specify mode by user
    """
    excel_data = readExcelWithPandas(input_dir, sheet)
    detail_data = getDetailSheet(input_dir, sheet)
    header = getListHeader(excel_data)
    filter_data = getListDataColumn(excel_data, header[filter])
    if (mode == 0):
        splitTableDataToMultiFile(
            excel_data,
            filter_data,
            header[filter],
            output_dir)
    elif (mode == 1):
        splitTableDataToMultiSheet(
            excel_data,
            detail_data,
            filter_data,
            header[filter],
            output_dir,
            sheet)


def splitToMultiDataWithMode(mode):
    """
    read file excel and split
    Args:
        mode: split mode
    """
    param = inputSplitParam()
    input = param.get('inputDir')
    output = param.get('outputDir')
    filter = param.get('filter', 0)
    sheet = param.get('activeSheet', 0)
    splitDataWithMode(input, output, filter, sheet, mode)
    print("===========DONE===========")


def createTableData(rows_data):
    """
    create new table data
    Args:
        rows_data: origin data
    """
    column_need = [9, 8, None, 3, None, 5, 7, 10, 0]
    tmp_table = []
    tmp_table_value = []
    for i, row_data in enumerate(rows_data):
        if i == 147:
            for x in range(0, 13):
                print(f'row data: {row_data[x].value}')
        tmp_rơw = []
        tmp_rơw_value = []
        for j in column_need:
            if j is not None:
                tmp_rơw_value.append(row_data[j].value)
                tmp_rơw.append(row_data[j])
            else:
                tmp_rơw_value.append(None)
                tmp_rơw.append(None)
        tmp_table.append(tmp_rơw)
        tmp_table_value.append(tmp_rơw_value)
    return tmp_table


def createFileWithOpenpyxl(rows_data, output_dir):
    """
    create file with openxlsx
    Args:
        output_dir: specify out put folder
        data: out put data
    """
    column_need = [9, 8, None, 3, None, 5, 7, 10, 0]
    total = len(rows_data)
    print("Total: ", total)
    progresBar = tqdm(range(total), desc="Create Fille...")
    wb = openpyxl.Workbook()
    sheet = wb.active

    for i, row_data in enumerate(rows_data):
        for j, column in enumerate(column_need):
            sheet.cell(i + 8, 1).value = i + 1
            if column is not None:
                sheet.cell(i + 8, j + 2).value = row_data[column].value
            else:
                sheet.cell(i + 8, j + 2).value = None
        progresBar.update()
    file_name = os.path.join(output_dir, "Test.xlsx")
    wb.save(file_name)


def isRowEmpty(row, min_column, max_column, row_number):
    """
    check row is empty
    Args:
        row: row data
        min_column: min column
        max_column: max column
    return:
        isEmpty
    """
    for i in range(min_column, max_column):
        cell = row[i]
        value = cell.value
        if value is not None:
            return False
    return True


def isRowNeed(row, min_column, max_column, row_number):
    """
    check row is copy
    Args:
        input_dir: specify input file
        sheet: active sheet
    """
    BG_COLOR = 64
    FG_COLOR = "FFFFFF00"
    is_empty = isRowEmpty(row, min_column, max_column, row_number)
    if is_empty is True:
        return False

    isRow = True
    for i in range(min_column, max_column):
        cell = row[i]
        value = cell.value
        bg_color = cell.fill.bgColor.index
        fg_color = cell.fill.fgColor.index
        if value is None:
            continue
        if bg_color == BG_COLOR and fg_color == FG_COLOR:
            continue
        else:
            isRow = False
            break

    if isRow is True:
        for i in range(min_column, max_column):
            cell = row[i]
            if cell.value is not None:
                bg_color = cell.fill.bgColor.index
                fg_color = cell.fill.fgColor.index
                if bg_color != BG_COLOR or fg_color != FG_COLOR:
                    isRow = False
                    break
    return isRow


def createTableDataWithColorFilter(input_dir, sheet, output_dir):
    """
    read and create filter color
    Args:
        input_dir: specify input file
        sheet: active sheet
        output_dir: specify output dir
    """
    rows_need = []
    work_sheet = readExcelWithOpenpyxl(input_dir, sheet)
    total = 0
    for i in range(2, work_sheet.max_row):
        row = work_sheet[i]
        isRow = isRowNeed(row, 0, 13, i)
        if isRow is True:
            total += 1
            rows_need.append(row)
    rows_length = len(rows_need)
    print(f'rows_length = {rows_length}, total = {total}')
    createFileWithOpenpyxl(rows_need, output_dir)


def createFilterColorMode(mode):
    """
    read and create filter color
    Args:
        mode: selected mode
    """
    param = inputParamForCreateFilter()
    input = param.get('inputDir')
    output = param.get('outputDir')
    sheet = param.get('activeSheet', 0)
    createTableDataWithColorFilter(input, sheet, output)
    print("===========DONE===========")


def selectMode():
    """
    ask the user to specify mode
    """
    modes = [
        "Split to multiple file",
        "Split to multiple sheet",
        "check color Mode"]
    print("All mode:")
    i = 0
    for mode in modes:
        print(f'   {i}: {mode}')
        i += 1
    print("==========================")
    input_mode = input("select mode: ")
    return input_mode


def run(mode):
    """
    run specify mode
    Args:
        mode: selected mode
    """
    if (mode == 0):
        splitToMultiDataWithMode(mode)
    elif (mode == 1):
        splitToMultiDataWithMode(mode)
    elif (mode == 2):
        createFilterColorMode(mode)
    else:
        print("mode dont exis")


if __name__ == "__main__":
    try:
        mode = selectMode()
        run(int(mode))
    except Exception as exc:
        print(f"failed to run split excel ({exc.args}).")
        print(f'{exc}')
